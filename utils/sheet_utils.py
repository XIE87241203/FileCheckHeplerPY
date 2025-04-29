from openpyxl import Workbook, load_workbook
import os

CONFIG_FILE_NAME = "config.xlsx"
CACHE_FILE_NAME = "cache.xlsx"
RESULT_FILE_NAME = "search_result.xlsx"
SEARCH_CONTENT_SHEET_NAME = "待搜索表"
SEARCH_RESULT_SHEET_NAME = "搜索结果"
DEL_FILE_SHEET_NAME = "待删除文件表"
FILE_WAREHOUSE_SHEET_NAME = "文件仓库"
ALL_FILE_NAME_SHEET_NAME = "仓库扫描结果"
HYPER_LINK_URL = "https://javdb.com/search?f=all&q="


# 检查并创建相关文件
def check_and_create_config_sheet():
    if os.path.exists(CONFIG_FILE_NAME):
        print("文件或目录存在，无需初始化配置表")
        return
    else:
        print("路径不存在，开始创建配置表")
        create_config_sheet()


def create_config_sheet():
    # 创建新工作簿（默认包含一个Sheet）
    config_book = Workbook()

    search_content_ws = config_book.active
    search_content_ws.title = SEARCH_CONTENT_SHEET_NAME
    # 待搜索文件名表
    search_content_headers = ["文件名"]
    search_content_ws.append(search_content_headers)
    # 文件仓库路径表
    warehouse_ws = config_book.create_sheet(title=FILE_WAREHOUSE_SHEET_NAME)
    warehouse_headers = ["文件仓库路径"]
    warehouse_ws.append(warehouse_headers)
    # 待删除文件路径表
    del_file_ws = config_book.create_sheet(title=DEL_FILE_SHEET_NAME)
    del_file_headers = ["待删除文件路径"]
    del_file_ws.append(del_file_headers)
    # 保存文件
    config_book.save(CONFIG_FILE_NAME)


# 读取仓库路径
def get_warehouse_path_list():
    path_list = []
    wb = load_workbook(CONFIG_FILE_NAME, read_only=True)
    # 获取仓库表
    sheet = wb[FILE_WAREHOUSE_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 仅返回值
        if any(cell is not None for cell in row):  # 过滤全空行
            path_list.append(row[0])
    # 处理数据...
    wb.close()  # 必须手动关闭
    return path_list

# 读取仓库路径
def get_del_path_list():
    path_list = []
    wb = load_workbook(CONFIG_FILE_NAME, read_only=True)
    # 获取仓库表
    sheet = wb[FILE_WAREHOUSE_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 仅返回值
        if any(cell is not None for cell in row):  # 过滤全空行
            path_list.append(row[0])
    # 处理数据...
    wb.close()  # 必须手动关闭
    return path_list

# 创建索引表
def create_cache_table(video_map):
    if os.path.exists(CACHE_FILE_NAME):
        os.remove(CACHE_FILE_NAME)  # 删除文件
        print("删除旧的索引文件")

    # 创建新工作簿（默认包含一个Sheet）
    cache_book = Workbook()

    files_ws = cache_book.active
    files_ws.title = ALL_FILE_NAME_SHEET_NAME
    # 待搜索文件名表
    files_ws_headers = ["文件名", "文件仓库路径", "文件数"]
    files_ws.append(files_ws_headers)
    # 文件仓库路径表
    warehouse_ws = cache_book.create_sheet(title=FILE_WAREHOUSE_SHEET_NAME)
    warehouse_headers = ["文件仓库路径"]
    warehouse_ws.append(warehouse_headers)

    # 遍历字典，写入数据
    for row, (key, value) in enumerate(
        video_map.items(), start=2
    ):  # start=2 从第2行开始
        path = "\n".join(value)
        files_ws.cell(row=row, column=1, value=key)  # A列写入 key
        files_ws.cell(row=row, column=2, value=path)  # B列写入 value
        files_ws.cell(row=row, column=3, value=len(value))  # B列写入相同文件数

    # 保存文件
    cache_book.save(CACHE_FILE_NAME)


def get_cache_map():
    cache_map = {}
    workbook = load_workbook(CACHE_FILE_NAME)
    sheet = workbook[ALL_FILE_NAME_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
        if row[0]:  # 确保第一列不为空
            cache_map[row[0]] = row[1]  # 第一列为key，第二列为value
    return cache_map


# 获取带搜索列表
def get_search_content_list():
    result = []
    workbook = load_workbook(CONFIG_FILE_NAME)
    sheet = workbook[SEARCH_CONTENT_SHEET_NAME]
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始
        if row[0]:  # 确保第一列不为空
            result.append(row[0])
    return result


def create_result_sheet(match_map, mismatch_list):
    if os.path.exists(RESULT_FILE_NAME):
        os.remove(RESULT_FILE_NAME)  # 删除文件
        print("删除旧的搜索结果文件")

    # 创建新工作簿（默认包含一个Sheet）
    result_book = Workbook()

    search_result_ws = result_book.active
    search_result_ws.title = SEARCH_RESULT_SHEET_NAME
    # 搜索结果标题
    search_result_ws_headers = ["文件名", "文件路径", "网站链接"]
    search_result_ws.append(search_result_ws_headers)
    index = 0
    # 遍历字典，写入搜索到的数据
    for row, (key, value) in enumerate(
        match_map.items(), start=2
    ):  # start=2 从第2行开始
        search_result_ws.cell(row=row, column=1, value=key)  # A列写入 key
        search_result_ws.cell(row=row, column=2, value=value)  # B列写入 value
        link_cell = search_result_ws.cell(
            row=row, column=3, value="访问网址"
        )  # C列写入 链接
        link_cell.hyperlink = get_hyperlink(key)
        link_cell.style = "Hyperlink"
        index = row + 1
    # 遍历未搜索到的结果，写入数据
    for mismatch in mismatch_list:
        search_result_ws.cell(row=index, column=1, value=mismatch)  # A列写入 名字
        link_cell = search_result_ws.cell(
            row=index, column=3, value="访问网址"
        )  # C列写入 链接
        link_cell.hyperlink = get_hyperlink(mismatch)
        link_cell.style = "Hyperlink"
        index += 1
    # 保存文件
    result_book.save(RESULT_FILE_NAME)


def get_hyperlink(item_name):
    return f"{HYPER_LINK_URL}{item_name}"
